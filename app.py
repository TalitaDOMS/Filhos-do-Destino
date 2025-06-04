from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
ARQUIVO = 'fichas.xlsx'

if not os.path.exists(ARQUIVO):
    wb = Workbook()
    wb.title = 'Resumo do RPG'
    ws = wb.active
    ws.append(['Nome do personagem','Nome do jogador','Parente Divino', 'Armamento','Humanidade', 'Força','Destreza', 'Vigor','Inteligência', 'Carisma','Aparência','Vida','Estamina','Áspis', 'Foco - Atributo'])
    wb.save(ARQUIVO)

@app.route('/')
def teste():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    return render_template('index.html')

@app.route('/fichas')
def fichas():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    dados = list(ws.iter_rows(min_row=2, values_only=True))
    return render_template('fichas.html', dados = dados)

@app.route('/salvar', methods=['POST'])
def salvar():
    nomeJogador = request.form['nomeJogador']
    nomePersonagem = request.form['nomePersonagem']
    parenteDivino = request.form['parenteDivino']
    armamento = request.form['armamento']
    humanidade = request.form['humanidade']
    forca = int(request.form['forca'])
    destreza = int(request.form['destreza'])
    vigor = int(request.form['vigor'])
    inteligencia = int(request.form['inteligencia'])
    carisma = int(request.form['carisma'])
    aparencia = int(request.form['aparencia'])
    vida = vigor * 2
    estamina = destreza * vigor
    focoAtributo = None
    aspis = 10

    # Calculo modificador parente divino
    if parenteDivino == 'Zeus':
        vida +=28
        estamina +=12
        focoAtributo = 'Vigor'
        aspis += vigor
    elif parenteDivino == 'Poseidon':
        vida +=12
        estamina +=25
        focoAtributo = 'Força'
        aspis += forca
    elif parenteDivino == 'Hades':
        vida +=8
        estamina +=35
        focoAtributo = 'Destreza'
        aspis += destreza

    elif parenteDivino == 'Atena':
        vida +=14
        estamina +=14
        focoAtributo = 'Inteligência'
        aspis += inteligencia
    elif parenteDivino == 'Ares':
        vida +=35
        estamina +=8
        focoAtributo = 'Força'
        aspis += forca
    elif parenteDivino == 'Hermes':
        vida +=8
        estamina +=45
        focoAtributo = 'Destreza'
        aspis += destreza
    elif parenteDivino == 'Apolo':
        vida +=14
        estamina +=18
        focoAtributo = 'Carisma'
        aspis += carisma
    elif parenteDivino == 'Afrodite':
        vida +=10
        estamina +=30
        focoAtributo = 'Aparência'
        aspis += aparencia
    elif parenteDivino == 'Hefesto':
        vida +=16
        estamina +=25
        focoAtributo = 'Vigor'
        aspis += vigor
    elif parenteDivino == 'Dionísio':
        vida +=15
        estamina += 20
        focoAtributo = 'Aparência'
        aspis += aparencia
    elif parenteDivino == 'Demeter':
        vida +=20
        estamina +=25
        focoAtributo = 'Inteligência'
        aspis += inteligencia
    elif parenteDivino == 'Nike':
        vida +=15
        estamina +=30
        focoAtributo = 'Vigor'
        aspis += vigor

    # Calculo modificador de armamento
    if armamento == 'Espadachim':
        vida +=5
        estamina +=4
    elif armamento == 'Atirador':
        vida +=3
        estamina +=6
    elif armamento == 'Lutador':
        vida +=8
        estamina +=2

    # Calculo modificador de humanidade
    if humanidade == 'Empático':
       estamina +=12
    elif humanidade == 'Monstruoso':
        vida += 12

    wb = load_workbook(ARQUIVO)
    ws = wb.active
    ws.append([nomePersonagem,nomeJogador,parenteDivino,armamento,humanidade,forca,destreza,vigor,inteligencia,carisma,aparencia,vida,estamina,aspis,focoAtributo])
    wb.save(ARQUIVO)
    dados = list(ws.iter_rows(min_row=2, values_only=True))
    return render_template('fichas.html', dados=dados)

# Rodar o aplicativo
if __name__ == '__main__':
    app.run(debug=True)